from pathlib import Path
from typing import Protocol

from langchain_community.document_loaders import (
    CSVLoader,
    Docx2txtLoader,
    PyPDFLoader,
    TextLoader,
)
from langchain_core.documents import Document


class TextExtractor(Protocol):
    def extract(self, path: Path) -> str: ...


class LangChainTextExtractor:
    """Extract plain text from PDF, TXT, DOCX, CSV, XLSX."""

    SUPPORTED = {"pdf", "txt", "docx", "csv", "xlsx"}

    def extract(self, path: Path) -> str:
        suffix = path.suffix.lower().lstrip(".")
        if suffix not in self.SUPPORTED:
            raise ValueError(f"Unsupported file type: {suffix}")

        if suffix == "pdf":
            loader = PyPDFLoader(str(path))
        elif suffix == "txt":
            loader = TextLoader(str(path), encoding="utf-8")
        elif suffix == "docx":
            loader = Docx2txtLoader(str(path))
        elif suffix == "csv":
            loader = CSVLoader(str(path))
        else:
            # Prefer openpyxl-based extraction without unstructured deps
            return self._extract_xlsx(path)

        docs: list[Document] = loader.load()
        text = "\n\n".join(d.page_content for d in docs if d.page_content)
        if not text.strip():
            raise ValueError("No text could be extracted from the file")
        return text

    @staticmethod
    def _extract_xlsx(path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "openpyxl is required for XLSX extraction"
            ) from exc

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            parts: list[str] = []
            for sheet in wb.worksheets:
                parts.append(f"# Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    cells = [
                        str(c).strip()
                        for c in row
                        if c is not None and str(c).strip()
                    ]
                    if cells:
                        parts.append(" | ".join(cells))
            text = "\n".join(parts)
        finally:
            wb.close()

        if not text.strip():
            raise ValueError("No text could be extracted from the spreadsheet")
        return text
